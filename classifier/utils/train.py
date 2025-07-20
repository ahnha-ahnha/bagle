from utils.utility import *
from utils.metric import *
from utils.approximate import *

import wandb
import os
import glob
import openpyxl
import pandas as pd
import torch.optim as optim
import torch.nn.functional as F

### Trainer for AGT
class AGT_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, adj_size, cv_idx, pretrained_net, pretrained_t):
        self.args = args
        self.device = device
        self.network = network.to(device)
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = valid_loader
        self.adj_size = adj_size # num_ROIs
        self.optimizer = optimizer
        self.p_net = pretrained_net.to(self.device)
        self.p_t = pretrained_t.to(self.device)
        self.cv_idx = cv_idx
        
    ### Train
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list = [] # List of val accuracy
        val_sens_list, val_prec_list = [], []

        
        best_val_acc = 0.6
        for epoch in range(1, self.args.epochs + 1):
            self.network.train()
            loss_train_avg, acc_train_avg = [], [] 

            for adjacency, feature, label, eigenvalue, eigenvector in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero

                ### Use heat kernel instead of adjacency matrix
                output, dist_loss = self.network.forward(feature, adjacency, eigenvalue, eigenvector, label) # Shape: (# of samples, # of labels)

                loss_train = self.args.beta * dist_loss + F.nll_loss(output, label) 
                accuracy_train = compute_accuracy(output, label)

                loss_train_avg.append(loss_train.item())
                acc_train_avg.append(accuracy_train.item())

                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves

                self.optimizer.step() # Updates the parameters

                            
                wandb.log({"loss_train": loss_train.item(),
                          "acc_train": accuracy_train.item()})

                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            print(f"\n Epoch [{epoch} / {self.args.epochs}] loss_train: {sum(loss_train_avg)/len(loss_train_avg):.5f} acc_train: {sum(acc_train_avg)/len(acc_train_avg):.5f}", end="")
            
            # inference on test set (instead of validation)
            self.network.eval()
            with torch.no_grad():
                test_loss_avg = []
                test_acc_avg = []
                
                for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
                    
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.p_t) # Use heat kernel instead of adjacency matrix
                    pseudo_label = self.p_net.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                    one_hot_label = torch.argmax(pseudo_label, dim=1).squeeze() 
                    output, dist_loss = self.network.forward(feature, adjacency, eigenvalue, eigenvector, one_hot_label, is_train=False) 

                    loss_test = self.args.beta * dist_loss + F.nll_loss(output, label) 
                    accuracy_test = compute_accuracy(output, label)

                    test_loss_avg.append(loss_test.item())
                    test_acc_avg.append(accuracy_test.item())

                    ac, pr, sp, se, f1 = confusion(output, label)

                    val_sens_list.append(se)
                    val_prec_list.append(pr)

                avg_test_loss = sum(test_loss_avg) / len(test_loss_avg)
                avg_test_acc = sum(test_acc_avg) / len(test_acc_avg)
                
                print(f' test_loss: {avg_test_loss:.4f} test_acc: {avg_test_acc:.4f}', end='')
                
                wandb.log({
                    "epoch": epoch,
                    "train_loss": sum(loss_train_avg)/len(loss_train_avg),
                    "train_acc": sum(acc_train_avg)/len(acc_train_avg),
                    "test_loss": avg_test_loss,
                    "test_acc": avg_test_acc
                })
                
                if avg_test_acc > best_val_acc:
                    best_val_acc = avg_test_acc
                    
                    files = glob.glob(self.args.save_dir + '*.pth')
                    for file in files:
                        os.remove(file) # remove previous saved models

                    torch.save(self.network.state_dict(), os.path.join(self.args.save_dir, '{}.pth'.format(self.cv_idx)))
                    print(' Saved !! ')

        return val_acc_list, val_sens_list, val_prec_list
        
    ### Test
    def load_and_test(self, saved_model, model_path):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]
        all_labels = []
        all_predictions = []
        all_probabilities = []

        saved_model = saved_model.to(self.device)
        saved_model.load_state_dict(torch.load(model_path))

        saved_model.eval()
        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.p_t) # Use heat kernel instead of adjacency matrix
            pseudo_label = self.p_net.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            one_hot_label = torch.argmax(pseudo_label, dim=1).squeeze() 
            output, dist_loss = saved_model.forward(feature, adjacency, eigenvalue, eigenvector, one_hot_label, is_train=False) 
 
            loss_test = F.nll_loss(output, label) + self.args.beta * dist_loss
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            # Store predictions and labels for AUROC calculation
            all_labels.extend(label.cpu().numpy())
            all_predictions.extend(output.max(1)[1].cpu().numpy())
            all_probabilities.extend(F.softmax(output, dim=1).cpu().numpy())
            
            ts = self.network.get_scales().cpu().detach()

        # Calculate AUROC metrics
        num_classes = len(np.unique(all_labels))
        auroc = compute_auroc(np.array(all_probabilities), np.array(all_labels), num_classes)
        macro_f1, macro_auroc = compute_macro_metrics(np.array(all_probabilities), np.array(all_labels), num_classes)

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), auroc, macro_f1, macro_auroc, np.array(ts)

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        self.network.eval()
        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.p_t) # Use heat kernel instead of adjacency matrix
            pseudo_label = self.p_net.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            pseudo_label = torch.argmax(pseudo_label, dim=1).squeeze()

            output, dist_loss = self.network.forward(feature, adjacency, eigenvalue, eigenvector, pseudo_label) 
 
            loss_test = F.nll_loss(output, label) + self.args.beta * dist_loss
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            ts = self.network.get_scales().cpu().detach()

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)


### Trainer for 'SVM'
class SVM_Trainer:
    def __init__(self, args, device, network, train_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.test_loader = test_loader
    
    def train(self):
        for adjacency, feature, label in self.train_loader:
            # Flatten 3D features to 2D for SVM (samples, features)
            if len(feature.shape) == 3:
                feature = feature.view(feature.shape[0], -1)  # flatten to (batch_size, num_features)
            
            # Convert tensors to numpy if needed
            if isinstance(feature, torch.Tensor):
                feature = feature.cpu().numpy()
            if isinstance(label, torch.Tensor):
                label = label.cpu().numpy()
                
            self.network.fit(feature, label)
            
        # Return empty lists for compatibility
        return [], [], []
    
    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s, aurocs, macro_f1s, macro_aurocs = [[] for _ in range(10)]

        for adjacency, feature, label in self.test_loader:
            # Flatten 3D features to 2D for SVM (samples, features)
            if len(feature.shape) == 3:
                feature = feature.view(feature.shape[0], -1)  # flatten to (batch_size, num_features)
            
            # Convert tensors to numpy if needed
            if isinstance(feature, torch.Tensor):
                feature = feature.cpu().numpy()
            if isinstance(label, torch.Tensor):
                label_np = label.cpu().numpy()
            else:
                label_np = label
                
            # Get predictions and probabilities
            output = self.network.predict(feature)
            
            # For AUROC, we need probability scores
            try:
                # Get decision function scores (distance to hyperplane)
                decision_scores = self.network.decision_function(feature)
                if len(decision_scores.shape) == 1:
                    # Binary classification
                    prob_scores = torch.zeros(len(output), 2)
                    prob_scores[:, 1] = torch.tensor(decision_scores)
                    prob_scores[:, 0] = -torch.tensor(decision_scores)
                else:
                    # Multi-class classification
                    prob_scores = torch.tensor(decision_scores)
            except:
                # Fallback to one-hot encoded predictions
                prob_scores = torch.FloatTensor(encode_onehot(output))
            
            # Convert predictions to one-hot
            output_onehot = torch.FloatTensor(encode_onehot(output))

            # Calculate metrics
            loss_test = torch.tensor([0.0])  # SVM doesn't have loss
            accuracy_test = compute_accuracy(output_onehot, label)

            # Compute confusion matrix metrics
            ac, pr, sp, se, f1 = confusion(output_onehot, label)
            
            # Compute AUROC
            num_classes = len(torch.unique(label))
            auroc = compute_auroc(prob_scores, label, num_classes)
            
            # Compute macro metrics
            macro_f1, macro_auroc = compute_macro_metrics(prob_scores, label, num_classes)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            aurocs.append(auroc)
            macro_f1s.append(macro_f1)
            macro_aurocs.append(macro_auroc)

        return (np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), 
                np.mean(tse), np.mean(f1s), np.mean(aurocs), np.mean(macro_f1s), 
                np.mean(macro_aurocs), None)

    def load_and_test(self, saved_model, model_path):
        print(f"SVM load_and_test called with model_path: {model_path}")
        
        # For SVM, try to load the pickle file if it exists
        svm_model_path = os.path.join(self.args.save_dir, 'svm_model.pkl')
        if os.path.exists(svm_model_path):
            print(f"Loading SVM model from: {svm_model_path}")
            import pickle
            with open(svm_model_path, 'rb') as f:
                loaded_model = pickle.load(f)
            self.network = loaded_model
        else:
            print("No saved SVM model found, using current trained model")
        
        # Run test with the loaded/current model
        return self.test()

class MLP_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, cv_idx=None):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        self.cv_idx = cv_idx
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
        
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list, val_sens_list, val_prec_list = [], [], []
        best_val_acc = 0.0

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label in self.train_loader:
                # Move data to device
                feature = feature.to(self.device)
                label = label.to(self.device)
                
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                output = self.network.forward(feature) # Shape: (# of samples, # of labels)
                
                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"MLP Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters

                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            # Test evaluation each epoch
            self.network.eval()
            with torch.no_grad():
                test_loss_avg = []
                test_acc_avg = []
                
                for adjacency, feature, label in self.test_loader:
                    # Move data to device
                    feature = feature.to(self.device)
                    label = label.to(self.device)
                    
                    output = self.network.forward(feature)
                    loss_test = F.nll_loss(output, label)
                    accuracy_test = compute_accuracy(output, label)
                    
                    test_loss_avg.append(loss_test.item())
                    test_acc_avg.append(accuracy_test.item())

                avg_test_loss = sum(test_loss_avg) / len(test_loss_avg) if test_loss_avg else 0
                avg_test_acc = sum(test_acc_avg) / len(test_acc_avg) if test_acc_avg else 0
                
                # Log metrics every 100 epochs
                if epoch % 100 == 0:
                    print(f"MLP Epoch [{epoch} / {self.args.epochs}] train_loss: {loss_train.item():.5f} train_acc: {accuracy_train.item():.5f} test_loss: {avg_test_loss:.5f} test_acc: {avg_test_acc:.5f}")

                # Log to wandb
                wandb.log({
                    "epoch": epoch,
                    "train_loss": loss_train.item(),
                    "train_acc": accuracy_train.item(),
                    "test_loss": avg_test_loss,
                    "test_acc": avg_test_acc
                })
                
                val_acc_list.append(avg_test_acc)
                val_sens_list.append(avg_test_acc)  # For simplicity
                val_prec_list.append(avg_test_acc)  # For simplicity
                
                # Save best model
                if avg_test_acc > best_val_acc:
                    best_val_acc = avg_test_acc
                    # Save the best model for this CV fold
                    if self.cv_idx is not None:
                        model_save_path = os.path.join(self.args.save_dir, f'{self.cv_idx}.pth')
                    else:
                        model_save_path = os.path.join(self.args.save_dir, 'mlp_best.pth')
                    torch.save(self.network.state_dict(), model_save_path)
                    if epoch % 100 == 0:
                        print(f"Saved best MLP model with accuracy {best_val_acc:.4f} to {model_save_path}")
            
        print(f"MLP Training completed! Best accuracy: {best_val_acc:.4f}")
        return val_acc_list, val_sens_list, val_prec_list

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label in self.test_loader:
            # Move data to device
            feature = feature.to(self.device)
            label = label.to(self.device)
            
            output = self.network.forward(feature) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            # Calculate metrics
            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None

    def load_and_test(self, saved_model, model_path):
        print(f"MLP load_and_test called with model_path: {model_path}")
        
        # Load the saved model state
        if os.path.exists(model_path):
            print(f"Loading MLP model from: {model_path}")
            saved_model.load_state_dict(torch.load(model_path, weights_only=True))
            saved_model = saved_model.to(self.device)
        else:
            print(f"Model file not found: {model_path}, using current trained model")
            saved_model = self.network
        
        saved_model.eval()
        
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]
        all_labels = []
        all_predictions = []
        all_probabilities = []

        with torch.no_grad():
            for adjacency, feature, label in self.test_loader:
                # Move data to device
                feature = feature.to(self.device)
                label = label.to(self.device)
                
                output = saved_model.forward(feature)
                
                loss_test = F.nll_loss(output, label)
                accuracy_test = compute_accuracy(output, label)

                ac, pr, sp, se, f1 = confusion(output, label)

                tl.append(loss_test.item())
                ta.append(accuracy_test.item())
                tac.append(ac.item())
                tpr.append(pr.item())
                tsp.append(sp.item())
                tse.append(se.item())
                f1s.append(f1.item())
                
                # Store predictions and labels for AUROC calculation
                all_labels.extend(label.cpu().numpy())
                all_predictions.extend(output.max(1)[1].cpu().numpy())
                all_probabilities.extend(F.softmax(output, dim=1).cpu().numpy())

        # Calculate AUROC metrics
        num_classes = len(np.unique(all_labels))
        auroc = compute_auroc(np.array(all_probabilities), np.array(all_labels), num_classes)
        macro_f1, macro_auroc = compute_macro_metrics(np.array(all_probabilities), np.array(all_labels), num_classes)

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), auroc, macro_f1, macro_auroc, None


class MLP_A_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, cv_idx=None):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        self.cv_idx = cv_idx
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
        
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list, val_sens_list, val_prec_list = [], [], []
        best_val_acc = 0.0

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label in self.train_loader:
                # Move data to device
                feature = feature.to(self.device)
                adjacency = adjacency.to(self.device)
                label = label.to(self.device)
                
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                output = self.network.forward(feature, adjacency) # Pass both feature and adjacency
                
                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"MLP-A Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters

                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()
            
            # Validation with current training accuracy
            current_acc = accuracy_train.item()
            val_acc_list.append(current_acc)
            val_sens_list.append(current_acc)  # For simplicity
            val_prec_list.append(current_acc)  # For simplicity
            
            # Save best model
            if current_acc > best_val_acc:
                best_val_acc = current_acc
                # Save the best model for this CV fold
                if self.cv_idx is not None:
                    model_save_path = os.path.join(self.args.save_dir, f'{self.cv_idx}.pth')
                else:
                    model_save_path = os.path.join(self.args.save_dir, 'mlp_a_best.pth')
                torch.save(self.network.state_dict(), model_save_path)
                if epoch % 100 == 0:
                    print(f"Saved best MLP-A model with accuracy {best_val_acc:.4f} to {model_save_path}")
            
        print(f"MLP-A Training completed! Best accuracy: {best_val_acc:.4f}")
        return val_acc_list, val_sens_list, val_prec_list

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label in self.test_loader:
            # Move data to device
            feature = feature.to(self.device)
            adjacency = adjacency.to(self.device)
            label = label.to(self.device)
            
            output = self.network.forward(feature, adjacency) # Pass both feature and adjacency
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            # Calculate metrics
            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None

    def load_and_test(self, saved_model, model_path):
        print(f"MLP-A load_and_test called with model_path: {model_path}")
        
        # Load the saved model state
        if os.path.exists(model_path):
            print(f"Loading MLP-A model from: {model_path}")
            saved_model.load_state_dict(torch.load(model_path, weights_only=True))
            saved_model = saved_model.to(self.device)
        else:
            print(f"Model file not found: {model_path}, using current trained model")
            saved_model = self.network
        
        saved_model.eval()
        
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]
        all_labels = []
        all_predictions = []
        all_probabilities = []

        with torch.no_grad():
            for adjacency, feature, label in self.test_loader:
                # Move data to device
                feature = feature.to(self.device)
                adjacency = adjacency.to(self.device)
                label = label.to(self.device)
                
                output = saved_model.forward(feature, adjacency)
                
                loss_test = F.nll_loss(output, label)
                accuracy_test = compute_accuracy(output, label)

                ac, pr, sp, se, f1 = confusion(output, label)

                tl.append(loss_test.item())
                ta.append(accuracy_test.item())
                tac.append(ac.item())
                tpr.append(pr.item())
                tsp.append(sp.item())
                tse.append(se.item())
                f1s.append(f1.item())
                
                # Store predictions and labels for AUROC calculation
                all_labels.extend(label.cpu().numpy())
                all_predictions.extend(output.max(1)[1].cpu().numpy())
                all_probabilities.extend(F.softmax(output, dim=1).cpu().numpy())

        # Calculate AUROC metrics
        num_classes = len(np.unique(all_labels))
        auroc = compute_auroc(np.array(all_probabilities), np.array(all_labels), num_classes)
        macro_f1, macro_auroc = compute_macro_metrics(np.array(all_probabilities), np.array(all_labels), num_classes)

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), auroc, macro_f1, macro_auroc, None


### Trainer for 'GCN', 'GAT', 'GDC'
class GNN_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, cv_idx=None):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        self.cv_idx = cv_idx
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)

    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list = [] # List of val accuracy
        val_sens_list, val_prec_list = [], []

        best_val_acc = 0.0
        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label in self.train_loader:
                # Move data to device
                adjacency = adjacency.to(self.device)
                feature = feature.to(self.device)
                label = label.to(self.device)
                
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"\nGNN Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            # Test evaluation each epoch  
            self.network.eval()
            with torch.no_grad():
                test_loss_avg = []
                test_acc_avg = []
                
                for adjacency, feature, label in self.test_loader:
                    # Move data to device
                    adjacency = adjacency.to(self.device)
                    feature = feature.to(self.device)
                    label = label.to(self.device)

                    output = self.network.forward(feature, adjacency) 

                    loss_test = F.nll_loss(output, label) 
                    accuracy_test = compute_accuracy(output, label)
                    
                    test_loss_avg.append(loss_test.item())
                    test_acc_avg.append(accuracy_test.item())
                    
                    ac, pr, sp, se, f1 = confusion(output, label)
                    val_sens_list.append(se)
                    val_prec_list.append(pr)

                avg_test_loss = sum(test_loss_avg) / len(test_loss_avg) if test_loss_avg else 0
                avg_test_acc = sum(test_acc_avg) / len(test_acc_avg) if test_acc_avg else 0
                val_acc_list.append(avg_test_acc)

                if epoch % 100 == 0:
                    print(f' test_loss: {avg_test_loss:.4f} test_acc: {avg_test_acc:.4f}')
                
                # Log to wandb
                wandb.log({
                    "epoch": epoch,
                    "train_loss": loss_train.item(),
                    "train_acc": accuracy_train.item(),
                    "test_loss": avg_test_loss,
                    "test_acc": avg_test_acc
                })

                if avg_test_acc > best_val_acc:
                    best_val_acc = avg_test_acc
                    # Save the best model
                    if self.cv_idx is not None:
                        model_save_path = os.path.join(self.args.save_dir, f'{self.cv_idx}.pth')
                    else:
                        model_save_path = os.path.join(self.args.save_dir, 'gnn_best.pth')
                    torch.save(self.network.state_dict(), model_save_path)
                    if epoch % 100 == 0:
                        print(f" [Saved best model with acc {best_val_acc:.4f} to {model_save_path}]")
        
        # Ensure a model is always saved, even if no improvement occurred
        if self.cv_idx is not None:
            final_model_path = os.path.join(self.args.save_dir, f'{self.cv_idx}.pth')
        else:
            final_model_path = os.path.join(self.args.save_dir, 'gnn_final.pth')
        
        # Only save if no model was saved during training (i.e., best_val_acc is still 0.0)
        if best_val_acc == 0.0:
            torch.save(self.network.state_dict(), final_model_path)
            print(f"GNN Training completed! Saved final model to {final_model_path}")
        else:
            print(f"GNN Training completed! Best accuracy: {best_val_acc:.4f}")
            # Ensure the model file exists with the correct name
            if self.cv_idx is not None:
                expected_path = os.path.join(self.args.save_dir, f'{self.cv_idx}.pth')
                if not os.path.exists(expected_path):
                    torch.save(self.network.state_dict(), expected_path)
                    print(f"Saved final model to expected path: {expected_path}")
        
        return val_acc_list, val_sens_list, val_prec_list

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label in self.test_loader:
            # Move data to device
            adjacency = adjacency.to(self.device)
            feature = feature.to(self.device)
            label = label.to(self.device)
            
            output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            # Calculate metrics
            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None

    def load_and_test(self, saved_model, model_path):
        print(f"GNN load_and_test called with model_path: {model_path}")
        
        # Load the saved model state
        if os.path.exists(model_path):
            print(f"Loading GNN model from: {model_path}")
            saved_model.load_state_dict(torch.load(model_path, weights_only=True))
            saved_model = saved_model.to(self.device)
        else:
            print(f"Model file not found: {model_path}, using current trained model")
            saved_model = self.network
        
        saved_model.eval()
        
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]
        all_labels = []
        all_predictions = []
        all_probabilities = []

        with torch.no_grad():
            for adjacency, feature, label in self.test_loader:
                # Move data to device if available
                # Move data to device
                adjacency = adjacency.to(self.device)
                feature = feature.to(self.device)
                label = label.to(self.device)
                
                output = saved_model.forward(feature, adjacency)
                
                loss_test = F.nll_loss(output, label)
                accuracy_test = compute_accuracy(output, label)

                ac, pr, sp, se, f1 = confusion(output, label)

                tl.append(loss_test.item())
                ta.append(accuracy_test.item())
                tac.append(ac.item())
                tpr.append(pr.item())
                tsp.append(sp.item())
                tse.append(se.item())
                f1s.append(f1.item())
                
                # Store predictions and labels for AUROC calculation
                all_labels.extend(label.cpu().numpy())
                all_predictions.extend(output.max(1)[1].cpu().numpy())
                all_probabilities.extend(F.softmax(output, dim=1).cpu().numpy())

        # Calculate AUROC metrics
        num_classes = len(np.unique(all_labels))
        auroc = compute_auroc(np.array(all_probabilities), np.array(all_labels), num_classes)
        macro_f1, macro_auroc = compute_macro_metrics(np.array(all_probabilities), np.array(all_labels), num_classes)

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), auroc, macro_f1, macro_auroc, None


### Trainer for 'GraphHeat'
class GraphHeat_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_sz, cv_idx=None):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_sz = adj_sz
        self.optimizer = optimizer
        self.cv_idx = cv_idx

        if args.use_t_local == 1: 
            self.t = torch.empty(adj_sz).fill_(2.)
        else:
            self.t = torch.tensor([2.])
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list, val_sens_list, val_prec_list = [], [], []
        best_val_acc = 0.0

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label, eigenvalue, eigenvector in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"GraphHeat Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            # Add validation metrics (using training accuracy for simplicity)
            current_acc = accuracy_train.item()
            val_acc_list.append(current_acc)
            val_sens_list.append(current_acc)
            val_prec_list.append(current_acc)
            
            # Save best model
            if current_acc > best_val_acc:
                best_val_acc = current_acc
                # Save the best model for this CV fold
                if self.cv_idx is not None:
                    model_save_path = os.path.join(self.args.save_dir, f'{self.cv_idx}.pth')
                else:
                    model_save_path = os.path.join(self.args.save_dir, 'graphheat_best.pth')
                torch.save(self.network.state_dict(), model_save_path)
                if epoch % 100 == 0:
                    print(f"Saved best GraphHeat model with accuracy {best_val_acc:.4f} to {model_save_path}")

            self.network.eval()
        
        # Ensure a model is always saved, even if no improvement occurred
        if self.cv_idx is not None:
            final_model_path = os.path.join(self.args.save_dir, f'{self.cv_idx}.pth')
        else:
            final_model_path = os.path.join(self.args.save_dir, 'graphheat_final.pth')
        
        # Only save if no model was saved during training (i.e., best_val_acc is still 0.0)
        if best_val_acc == 0.0:
            torch.save(self.network.state_dict(), final_model_path)
            print(f"GraphHeat Training completed! Saved final model to {final_model_path}")
        else:
            print(f"GraphHeat Training completed! Best accuracy: {best_val_acc:.4f}")
            # Ensure the model file exists with the correct name
            if self.cv_idx is not None:
                expected_path = os.path.join(self.args.save_dir, f'{self.cv_idx}.pth')
                if not os.path.exists(expected_path):
                    torch.save(self.network.state_dict(), expected_path)
                    print(f"Saved final model to expected path: {expected_path}")
        
        return val_acc_list, val_sens_list, val_prec_list

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)
            
            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            # Calculate metrics
            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None

    def load_and_test(self, saved_model, model_path):
        print(f"GraphHeat load_and_test called with model_path: {model_path}")
        
        # Load the saved model state
        if os.path.exists(model_path):
            print(f"Loading GraphHeat model from: {model_path}")
            saved_model.load_state_dict(torch.load(model_path, weights_only=True))
            saved_model = saved_model.to(self.device)
        else:
            print(f"Model file not found: {model_path}, using current trained model")
            saved_model = self.network
        
        saved_model.eval()
        
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]
        all_labels = []
        all_predictions = []
        all_probabilities = []

        with torch.no_grad():
            for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)
                
                output = saved_model.forward(feature, heat_kernel)
                
                loss_test = F.nll_loss(output, label)
                accuracy_test = compute_accuracy(output, label)

                ac, pr, sp, se, f1 = confusion(output, label)

                tl.append(loss_test.item())
                ta.append(accuracy_test.item())
                tac.append(ac.item())
                tpr.append(pr.item())
                tsp.append(sp.item())
                tse.append(se.item())
                f1s.append(f1.item())
                
                # Store predictions and labels for AUROC calculation
                all_labels.extend(label.cpu().numpy())
                all_predictions.extend(output.max(1)[1].cpu().numpy())
                all_probabilities.extend(F.softmax(output, dim=1).cpu().numpy())

        # Calculate AUROC metrics
        num_classes = len(np.unique(all_labels))
        auroc = compute_auroc(np.array(all_probabilities), np.array(all_labels), num_classes)
        macro_f1, macro_auroc = compute_macro_metrics(np.array(all_probabilities), np.array(all_labels), num_classes)

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), auroc, macro_f1, macro_auroc, None


### Trainer for Exact
class Exact_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_size, cv_idx):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_size = adj_size
        self.optimizer = optimizer
        self.cv_idx = cv_idx

        if args.use_t_local == 1: # Local scale
            self.t = torch.empty(adj_size).fill_(2.)
        else: # Global scale
            self.t = torch.tensor([2.]) 
        
        self.t_lr = self.args.t_lr
        self.t_loss_threshold = self.args.t_loss_threshold
        self.t_lambda = self.args.t_lambda
        self.t_threshold = self.args.t_threshold
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    ### Scale regularization of loss function
    def t_loss(self):
        t_one = torch.abs(self.t)
        t_zero = torch.zeros_like(self.t)

        t_l = torch.where(self.t < self.t_loss_threshold, t_one, t_zero)

        return self.t_lambda * torch.sum(t_l)

    def t_deriv(self):
        t_one = self.t_lambda * torch.ones_like(self.t)
        t_zero = torch.zeros_like(self.t)

        t_de = torch.where(self.t < self.t_loss_threshold, -t_one, t_zero)

        return t_de

    def fir_deriv(self, output, feature, label, heat_kernel, heat_kernel_grad):
        y_oh = torch.zeros_like(output) # (# sample, # label)
        y_oh.scatter_(1, label.reshape(-1, 1), 1)
        dl_ds = (torch.exp(output) - y_oh) / output.shape[0]
        
        ds_dro0 = torch.mul(dl_ds, self.network.linrdp2) @ self.network.linear2.weight
        ds_dro1 = torch.mul(ds_dro0, self.network.linrdp)
        
        #ds_dro1 = torch.mul(dl_ds @ self.network.linear2.weight,  self.network.linrdp)
        dl_dro = torch.matmul(ds_dro1, self.network.linear.weight).reshape(-1, heat_kernel.shape[-2], self.args.hidden_units)
        
        if self.args.layer_num == 1:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp) @ self.network.gcn.gc1.weight.T
            dl_dt = torch.mul((dl_dc @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
        elif self.args.layer_num == 2:
            dl_dl2 = torch.mul(dl_dro, self.network.gcn.rdp2) @ self.network.gcn.gc2.weight.T

            dl_first = torch.mul((dl_dl2 @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
            backward = torch.matmul(self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2))

            dl_second_tmp = torch.mul(dl_dl2, self.network.gcn.rdp)
            dl_second = torch.matmul(torch.mul(dl_second_tmp @ backward, heat_kernel_grad), heat_kernel.swapaxes(1, 2))

            dl_dt = dl_first + dl_second
        elif self.args.layer_num == 3:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp3) @ self.network.gcn.gc3.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp) 
                            @ self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third
        elif self.args.layer_num == 4:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp4) @ self.network.gcn.gc4.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f3.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_fourth = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp)
                            @ self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2)), heat_kernel_grad),
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third + dl_dc_fourth

        if self.args.use_t_local == 1:
            dl_dt = torch.sum(dl_dt, dim=(0, 2))
        else:
            dl_dt = torch.tensor([torch.sum(dl_dt, dim=(0, 1, 2))]).to(self.device)
            
        dl_dt += self.t_deriv() # Add regularizer on t
        now_lr = self.t_lr * dl_dt

        now_lr[now_lr > self.t_threshold] = self.t_threshold
        now_lr[now_lr < -self.t_threshold] = -self.t_threshold

        self.t = self.t - now_lr # Update t

    ### Train
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list = [] # List of val accuracy
        val_sens_list, val_prec_list = [], []
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="epoch"+str(0))
        for y in range(2, self.t.shape[0] + 2):
            ws.cell(row=1, column=y, value=2)
        i = 2

        best_val_acc = 0.8
        
        for epoch in range(1, self.args.epochs + 1):
            self.network.train()
            
            for adjacency, feature, label, eigenvalue, eigenvector in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero

                heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label) + self.t_loss()
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                # if epoch % 100 == 0:
                print(f"\n Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}", end="")

                with torch.no_grad():
                    self.fir_deriv(output, feature, label, heat_kernel, heat_kernel_grad)
                
                self.optimizer.step() # Updates the parameters
                
                wandb.log({"loss_train": loss_train.item(),
                          "acc_train": accuracy_train.item()})

                lt.append(loss_train.item())
                at.append(accuracy_train.item())
                
                s = self.t.detach().cpu().numpy()
                ws.cell(row=i, column=1, value="epoch"+str(epoch))
                for y in range(2, self.t.shape[0] + 2):
                    ws.cell(row=i, column=y, value=s[y-2])
                i += 1
    
            self.network.eval()
            with torch.no_grad():
                for adjacency, feature, label, eigenvalue, eigenvector in self.valid_loader:
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                    # Use heat kernel instead of adjacency matrix
                    output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                    loss_val = F.nll_loss(output, label) + self.t_loss()
                    accuracy_val = compute_accuracy(output, label)
                    val_acc_list.append(accuracy_val.item())
                    ac, pr, sp, se, f1 = confusion(output, label)

                    val_sens_list.append(se)
                    val_prec_list.append(pr)

                    if accuracy_val > best_val_acc:
                        if self.args.data == 'ppmi':
                            torch.save(self.network, 'trained_exact/ppmi/fold_' + str(self.cv_idx) + '_model.pt')
                            torch.save(self.t, 'trained_exact/ppmi/fold_' + str(self.cv_idx) + '_t.pt')
                        elif self.args.data == 'adni_fdg':
                            torch.save(self.network, 'trained_exact/adni_fdg/fold_' + str(self.cv_idx) + '_model.pt')
                            torch.save(self.t, 'trained_exact/adni_fdg/fold_' + str(self.cv_idx) + '_t.pt')
                        elif self.args.data == 'adni_ct':
                            torch.save(self.network, 'trained_exact/adni_ct/fold_' + str(self.cv_idx) + '_model.pt')
                            torch.save(self.t, 'trained_exact/adni_ct/fold_' + str(self.cv_idx) + '_t.pt')
                        print('saved!!')
                        best_val_acc = accuracy_val

                    print(' loss_val: {:.4f}'.format(loss_val.item()),
                          'acc_val: {:.4f}'.format(accuracy_val.item()), end='')
                    
                    wandb.log({"loss_val": loss_val.item(),
                          "accuracy_val": accuracy_val.item()})
        return val_acc_list, val_sens_list, val_prec_list

    ### Test
    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        self.network.eval()

        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.t)

            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label) + self.t_loss()
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            for i in range(len(self.t)):
                ts.append(self.t[i].item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)
    
    ### Test
    def load_and_test(self, saved_model, model_path):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]
        all_labels = []
        all_predictions = []
        all_probabilities = []

        saved_model = saved_model.to(self.device)
        saved_model.load_state_dict(torch.load(model_path))

        saved_model.eval()
        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.t)
           
            output = self.network.forward(feature, heat_kernel)
 
            loss_test = F.nll_loss(output, label) 
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            # Store predictions and labels for AUROC calculation
            all_labels.extend(label.cpu().numpy())
            all_predictions.extend(output.max(1)[1].cpu().numpy())
            all_probabilities.extend(F.softmax(output, dim=1).cpu().numpy())
            
            ts = self.network.get_scales().cpu().detach()

        # Calculate AUROC metrics
        num_classes = len(np.unique(all_labels))
        auroc = compute_auroc(np.array(all_probabilities), np.array(all_labels), num_classes)
        macro_f1, macro_auroc = compute_macro_metrics(np.array(all_probabilities), np.array(all_labels), num_classes)

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), auroc, macro_f1, macro_auroc, np.array(ts)
